"""
NCC Cadet Strength & Gender-Diversity Trend — Automated Reporting Pipeline
---------------------------------------------------------------------------
Ingests four REAL, differently-structured NCC data releases (a Lok Sabha reply
on cadet strength trend, a PIB gender breakdown, a PIB institutional-coverage
release, and a Ministry of Defence expansion announcement), reconciles them
into one consolidated metrics table, validates the figures against each other,
loads them into SQLite, runs SQL analysis, and auto-generates a formatted
Excel report — with every figure traceable to its real source.

Every number here is a genuine reported statistic (see sources/ and the
Sources sheet in the generated report for exact attribution). Nothing is
simulated.

Usage:
    python3 etl_ncc_report.py
"""
import os
import sqlite3
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

SRC_DIR = "sources"
DB_PATH = "output/ncc_metrics.db"
REPORT_PATH = "output/NCC_Cadet_Strength_Diversity_Report.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
SUB_FONT = Font(name=FONT, italic=True, size=10, color="595959")
NORM = Font(name=FONT, size=10)
FLAG_FONT = Font(name=FONT, size=10, color="9C0006")
FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SOURCES = [
    ("lok_sabha_2025_cadet_strength_trend.csv", "Lok Sabha reply — total cadets enrolled & girls % (2021 vs 2025)", 2025),
    ("pib_2025_girls_wing_breakdown.csv", "PIB — girls cadets by Wing (Senior/Junior), as of 2025", 2025),
    ("pib_2025_institutional_coverage.csv", "PIB — institutions covered, pipeline, and training camps, as of 2025", 2025),
    ("mod_expansion_announcement.csv", "Ministry of Defence — sanctioned-strength expansion announcement", 2025),
]


def normalize_all():
    """Reads the four differently-shaped source files and reconciles them into
    one long-format table: metric_category, metric_name, year, value, unit, source_file."""
    records = []
    issues = []

    # --- Source 1: cadet strength trend (wide format, two years) ---
    df1 = pd.read_csv(os.path.join(SRC_DIR, "lok_sabha_2025_cadet_strength_trend.csv"))
    for _, row in df1.iterrows():
        records.append(dict(metric_category="Cadet Strength", metric_name="Total_Cadets_Enrolled",
                             year=int(row["Year"]), value=row["Total_Cadets_Enrolled"], unit="cadets",
                             source_file="lok_sabha_2025_cadet_strength_trend.csv"))
        records.append(dict(metric_category="Gender Diversity", metric_name="Girls_Percentage",
                             year=int(row["Year"]), value=row["Girls_Percentage"], unit="%",
                             source_file="lok_sabha_2025_cadet_strength_trend.csv"))

    # --- Source 2: girls by wing (needs summing to get a total for cross-check) ---
    df2 = pd.read_csv(os.path.join(SRC_DIR, "pib_2025_girls_wing_breakdown.csv"))
    girls_total_from_wings = df2["Girls_Count"].sum()
    for _, row in df2.iterrows():
        records.append(dict(metric_category="Gender Diversity", metric_name=f"Girls_Count_{row['Wing'].replace(' ', '_')}",
                             year=int(row["As_Of_Year"]), value=row["Girls_Count"], unit="cadets",
                             source_file="pib_2025_girls_wing_breakdown.csv"))
    records.append(dict(metric_category="Gender Diversity", metric_name="Girls_Count_Total_(summed_from_wings)",
                         year=2025, value=girls_total_from_wings, unit="cadets",
                         source_file="pib_2025_girls_wing_breakdown.csv"))

    # cross-validation: does wing-summed girls count match Total_Cadets * Girls_Percentage?
    total_2025 = df1.loc[df1["Year"] == 2025, "Total_Cadets_Enrolled"].iloc[0]
    pct_2025 = df1.loc[df1["Year"] == 2025, "Girls_Percentage"].iloc[0]
    implied_girls = round(total_2025 * pct_2025 / 100)
    diff = abs(implied_girls - girls_total_from_wings)
    if diff > 0:
        issues.append({
            "Issue": "Cross-source figure mismatch (not silently corrected)",
            "Detail": (f"Girls count implied by Total_Cadets x Girls_Percentage = {implied_girls:,}, "
                       f"but Senior+Junior Wing figures sum to {girls_total_from_wings:,}. "
                       f"Difference = {diff:,} (likely rounding in one of the two independently published figures)."),
            "Sources_Involved": "lok_sabha_2025_cadet_strength_trend.csv vs pib_2025_girls_wing_breakdown.csv",
        })

    # --- Source 3: institutional coverage (metric/value/unit, single year) ---
    df3 = pd.read_csv(os.path.join(SRC_DIR, "pib_2025_institutional_coverage.csv"))
    for _, row in df3.iterrows():
        records.append(dict(metric_category="Institutional Coverage", metric_name=row["Metric"],
                             year=int(row["As_Of_Year"]), value=row["Value"], unit=row["Unit"],
                             source_file="pib_2025_institutional_coverage.csv"))

    # --- Source 4: expansion plan (metric/value/unit/notes, no year column -> flagged & assigned) ---
    df4 = pd.read_csv(os.path.join(SRC_DIR, "mod_expansion_announcement.csv"))
    for _, row in df4.iterrows():
        records.append(dict(metric_category="Expansion Plan", metric_name=row["Metric"],
                             year=None, value=row["Value"], unit=row["Unit"],
                             source_file="mod_expansion_announcement.csv"))
    issues.append({
        "Issue": "Missing year/date field in source",
        "Detail": "mod_expansion_announcement.csv does not carry a publication year in the original release; "
                  "left as blank (year=None) rather than assumed, per source.",
        "Sources_Involved": "mod_expansion_announcement.csv",
    })

    metrics_df = pd.DataFrame(records)
    issues_df = pd.DataFrame(issues)
    return metrics_df, issues_df


def load_to_sqlite(metrics_df):
    os.makedirs("output", exist_ok=True)
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    metrics_df.to_sql("ncc_metrics", conn, index=False)
    return conn


def run_sql_analysis(conn):
    strength_trend = pd.read_sql_query("""
        SELECT year, value AS total_cadets_enrolled
        FROM ncc_metrics
        WHERE metric_name = 'Total_Cadets_Enrolled'
        ORDER BY year
    """, conn)

    girls_pct_trend = pd.read_sql_query("""
        SELECT year, value AS girls_percentage
        FROM ncc_metrics
        WHERE metric_name = 'Girls_Percentage'
        ORDER BY year
    """, conn)

    institutional = pd.read_sql_query("""
        SELECT metric_name, value, unit
        FROM ncc_metrics
        WHERE metric_category = 'Institutional Coverage'
        ORDER BY value DESC
    """, conn)

    expansion = pd.read_sql_query("""
        SELECT metric_name, value, unit
        FROM ncc_metrics
        WHERE metric_category = 'Expansion Plan'
    """, conn)

    # CAGR of cadet strength 2021 -> 2025 (n = 4 years)
    y0 = strength_trend.iloc[0]["total_cadets_enrolled"]
    y1 = strength_trend.iloc[-1]["total_cadets_enrolled"]
    n_years = int(strength_trend.iloc[-1]["year"]) - int(strength_trend.iloc[0]["year"])
    cagr = ((y1 / y0) ** (1 / n_years) - 1) * 100 if n_years else None

    return strength_trend, girls_pct_trend, institutional, expansion, cagr


def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_df(ws, df, start_row):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    style_header_row(ws, start_row, len(df.columns))
    for i, (_, rec) in enumerate(df.iterrows(), start=1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row + i, column=j, value=rec[col])
            cell.font = NORM
            cell.border = BORDER
    return start_row + len(df) + 1


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(strength_trend, girls_pct_trend, institutional, expansion, cagr, issues_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NCC Cadet Strength & Gender-Diversity Trend — Auto-Generated Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')} · Reconciled from 4 real government/press releases"
    ws["A2"].font = SUB_FONT

    r = 4
    ws.cell(row=r, column=1, value="Cadet Strength Trend").font = Font(name=FONT, bold=True, size=12)
    r += 1
    r = write_df(ws, strength_trend, r)
    st_start, st_end = r - len(strength_trend) - 1, r - 2

    r += 1
    ws.cell(row=r, column=1, value=f"CAGR (2021→2025): {cagr:.2f}% per year").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    r += 2

    ws.cell(row=r, column=1, value="Girls' Share of Cadet Strength (%)").font = Font(name=FONT, bold=True, size=12)
    r += 1
    r = write_df(ws, girls_pct_trend, r)
    gp_start, gp_end = r - len(girls_pct_trend) - 1, r - 2

    r += 1
    ws.cell(row=r, column=1, value="Institutional Coverage (as of 2025)").font = Font(name=FONT, bold=True, size=12)
    r += 1
    r = write_df(ws, institutional, r)
    inst_start, inst_end = r - len(institutional) - 1, r - 2

    r += 1
    ws.cell(row=r, column=1, value="Expansion Plan (announced targets)").font = Font(name=FONT, bold=True, size=12)
    r += 1
    r = write_df(ws, expansion, r)

    autofit(ws, [40, 20, 14])

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Total Cadets Enrolled: 2021 vs 2025"
    chart1.y_axis.title = "Cadets"
    data1 = Reference(ws, min_col=2, min_row=st_start, max_row=st_end)
    cats1 = Reference(ws, min_col=1, min_row=st_start + 1, max_row=st_end)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.height, chart1.width = 8, 14
    ws.add_chart(chart1, "E4")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Girls' Share of Cadet Strength: 2021 vs 2025 (%)"
    chart2.y_axis.title = "%"
    data2 = Reference(ws, min_col=2, min_row=gp_start, max_row=gp_end)
    cats2 = Reference(ws, min_col=1, min_row=gp_start + 1, max_row=gp_end)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height, chart2.width = 8, 14
    ws.add_chart(chart2, "E22")

    # --- Sources sheet ---
    ws2 = wb.create_sheet("Sources")
    ws2["A1"] = "Data Sources & Citations"
    ws2["A1"].font = TITLE_FONT
    src_hdr = 3
    for i, h in enumerate(["Source_File", "Description", "Year"]):
        ws2.cell(row=src_hdr, column=i + 1, value=h)
    style_header_row(ws2, src_hdr, 3)
    for i, (fname, desc, yr) in enumerate(SOURCES):
        rr = src_hdr + 1 + i
        ws2.cell(row=rr, column=1, value=fname).font = NORM
        ws2.cell(row=rr, column=2, value=desc).font = NORM
        ws2.cell(row=rr, column=3, value=yr).font = NORM
        for c in range(1, 4):
            ws2.cell(row=rr, column=c).border = BORDER
            ws2.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    autofit(ws2, [42, 70, 10])

    # --- Data quality log ---
    ws3 = wb.create_sheet("Data_Quality_Log")
    ws3["A1"] = "Cross-Source Validation & Data Quality Log"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = "Discrepancies between independently published figures are flagged here, not silently resolved."
    ws3["A2"].font = SUB_FONT
    if len(issues_df):
        rr = write_df(ws3, issues_df, 4)
        for row in range(5, rr):
            for c in range(1, 4):
                ws3.cell(row=row, column=c).fill = FLAG_FILL
                ws3.cell(row=row, column=c).font = FLAG_FONT
                ws3.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    autofit(ws3, [34, 70, 46])

    wb.save(REPORT_PATH)


def main():
    metrics_df, issues_df = normalize_all()
    conn = load_to_sqlite(metrics_df)
    strength_trend, girls_pct_trend, institutional, expansion, cagr = run_sql_analysis(conn)
    build_report(strength_trend, girls_pct_trend, institutional, expansion, cagr, issues_df)
    print(f"Metrics reconciled from {len(SOURCES)} real source files: {len(metrics_df)} rows")
    print(f"Cadet strength CAGR 2021->2025: {cagr:.2f}% per year")
    print(f"Cross-source issues flagged: {len(issues_df)}")
    print(f"Report written to: {REPORT_PATH}")


if __name__ == "__main__":
    main()
